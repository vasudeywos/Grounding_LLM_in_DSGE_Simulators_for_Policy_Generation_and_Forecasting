from __future__ import annotations

import json
from dataclasses import dataclass, replace
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from dsge_rl.config import ScenarioConfig
from dsge_rl.environment import DSGEPolicyEnvironment
from dsge_rl.snowdrop import SnowdropSimulator
from dsge_rl.simulator_registry import simulator_registry


DATASET_DIRECTORY = Path(__file__).resolve().parent.parent / "The Global Economic Impacts of the COVID-19 Pandemic"
SHOCKS_FILE = DATASET_DIRECTORY / "McKibbin & Fernando_2023_Shocks.xlsx"
RESULTS_FILE = DATASET_DIRECTORY / "McKibbin & Fernando_2023_Simulation Results.xlsx"


@dataclass(frozen=True)
class PandemicScenario:
    scenario: str
    region: str
    source_values: dict[str, float]
    gsw_shocks: dict[str, list[float]]
    approximate: bool


class PandemicDataset:
    def __init__(self, shocks_file: str | Path = SHOCKS_FILE, results_file: str | Path = RESULTS_FILE):
        self.shocks_file = Path(shocks_file)
        self.results_file = Path(results_file)

    def load_shocks(self, scenario: str, region: str) -> tuple[dict[str, float], bool]:
        scenario = scenario.upper()
        region = region.upper()
        approximate = scenario == "S06"
        source_scenario = "S01" if approximate else scenario
        if source_scenario not in {"S01", "S02", "S03", "S04", "S05"}:
            raise ValueError("scenario must be S01 through S06")
        workbook = load_workbook(self.shocks_file, read_only=True, data_only=True)
        sheet = workbook[f"S_{source_scenario}"]
        headers = [str(value) if value is not None else "" for value in next(sheet.iter_rows(values_only=True))]
        for row in sheet.iter_rows(values_only=True):
            if str(row[0]).upper() == region:
                values = {
                    header.split()[0]: float(value)
                    for header, value in zip(headers[2:], row[2:])
                    if header and value is not None
                }
                workbook.close()
                return values, approximate
        workbook.close()
        raise ValueError(f"Region {region} not found in {source_scenario}")

    def load_results(self, scenario: str, region: str, variable: str) -> dict[int, float]:
        workbook = load_workbook(self.results_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = {str(value): index for index, value in enumerate(next(rows))}
        result = {}
        for row in rows:
            if (
                str(row[headers["Variable"]]).upper() == variable.upper()
                and str(row[headers["Scenario"]]).upper() == scenario.upper()
                and str(row[headers["Region"]]).upper() == region.upper()
                and str(row[headers["Sector"]]).upper() == "NA"
            ):
                year = int(str(row[headers["Year"]]).removeprefix("Y"))
                result[year] = float(row[headers["Value"]])
        workbook.close()
        if not result:
            raise ValueError(f"No results for {scenario}, {region}, {variable}")
        return result


class GSWPandemicMapper:
    def __init__(self, periods: int = 28, active_quarters: int = 4, percentage_scale: float = 0.01):
        self.periods = periods
        self.active_quarters = active_quarters
        self.percentage_scale = percentage_scale

    def map(self, source: dict[str, float]) -> dict[str, list[float]]:
        productivity = self._mean(source, ("TFPP01", "TFPP02", "TFPP03", "TFPP04", "TFPP05", "TFPP06"))
        sector_risk = self._mean(source, ("RISEP01", "RISEP02", "RISEP03", "RISEP04", "RISEP05", "RISEP06"))
        mapped = {
            "ea": productivity,
            "ey": source.get("SHKC", 0.0),
            "eb": np.mean([source.get("RISH", 0.0), source.get("EXCR", 0.0), sector_risk]),
            "eg": source.get("GOVS", 0.0),
            "els": source.get("MORB", 0.0) + source.get("MORT", 0.0),
        }
        return {name: self._annual_path(float(value)) for name, value in mapped.items()}

    def _annual_path(self, value: float) -> list[float]:
        quarterly = value * self.percentage_scale / self.active_quarters
        return [quarterly] * self.active_quarters + [0.0] * (self.periods - self.active_quarters)

    def _mean(self, source: dict[str, float], names: tuple[str, ...]) -> float:
        values = [source[name] for name in names if name in source]
        return float(np.mean(values)) if values else 0.0


class PandemicValidator:
    def __init__(self, periods: int = 28, percentage_scale: float = 0.01):
        self.periods = periods
        self.percentage_scale = percentage_scale
        self.dataset = PandemicDataset()
        self.mapper = GSWPandemicMapper(periods=periods, percentage_scale=percentage_scale)
        self.gsw_config = simulator_registry(periods=periods, turns=4)["gsw"]

    def scenario(self, scenario: str, region: str) -> PandemicScenario:
        source, approximate = self.dataset.load_shocks(scenario, region)
        return PandemicScenario(scenario.upper(), region.upper(), source, self.mapper.map(source), approximate)

    def validate_simulator(self, scenario: str, region: str) -> dict:
        pandemic = self.scenario(scenario, region)
        simulator = SnowdropSimulator(self.gsw_config.model_path, self.periods)
        frame = simulator.simulate(pandemic.gsw_shocks, ("y", "pinf"))
        simulated = {
            "GDPR": self._annualize(frame["y"].to_numpy(dtype=float)),
            "INFL": self._annualize(frame["pinf"].to_numpy(dtype=float)),
        }
        comparisons = {}
        for variable, predicted in simulated.items():
            reference = self.dataset.load_results(pandemic.scenario, pandemic.region, variable)
            years = sorted(set(reference).intersection(predicted).difference({0}))
            actual = np.array([reference[year] for year in years], dtype=float)
            estimate = np.array([predicted[year] for year in years], dtype=float)
            comparisons[variable] = {
                "years": years,
                "g_cubed": actual.tolist(),
                "gsw": estimate.tolist(),
                "rmse": float(np.sqrt(np.mean(np.square(estimate - actual)))) if len(years) else None,
                "correlation": self._correlation(estimate, actual),
            }
        return {
            "scenario": pandemic.scenario,
            "region": pandemic.region,
            "scenario_6_approximation": pandemic.approximate,
            "percentage_scale": self.percentage_scale,
            "mapped_gsw_shocks": {key: values[:4] for key, values in pandemic.gsw_shocks.items()},
            "comparisons": comparisons,
        }

    def build_policy_environment(self, scenario: str, region: str) -> DSGEPolicyEnvironment:
        pandemic = self.scenario(scenario, region)
        discourse = self._discourse(pandemic)
        scenario_config = ScenarioConfig(
            name=f"COVID_{pandemic.scenario}_{pandemic.region}",
            shocks=pandemic.gsw_shocks,
            semantic_volatility=float(np.linalg.norm(list(pandemic.source_values.values()))) * self.percentage_scale,
            discourse=discourse,
        )
        config = replace(self.gsw_config, burn_in=0, scenarios=(scenario_config,))
        return DSGEPolicyEnvironment(config)

    def _annualize(self, values: np.ndarray) -> dict[int, float]:
        result = {0: 0.0}
        for year, start in enumerate(range(0, min(len(values), 24), 4), start=1):
            result[year] = float(np.mean(values[start : start + 4]))
        return result

    def _correlation(self, left: np.ndarray, right: np.ndarray) -> float | None:
        if len(left) < 2 or np.std(left) == 0 or np.std(right) == 0:
            return None
        return float(np.corrcoef(left, right)[0, 1])

    def _discourse(self, scenario: PandemicScenario) -> tuple[str, ...]:
        return (
            f"Before the pandemic shock, economic activity in {scenario.region} was near baseline.",
            f"Pandemic restrictions disrupted labor supply, productivity and consumption in {scenario.region}.",
            "Households reduced spending while firms reported shutdowns and rising risk premiums.",
            "Government support increased as output, employment and confidence remained under pressure.",
            "Restrictions gradually eased and discussion shifted toward recovery and persistent economic damage.",
        )


def write_validation_report(report: dict, path: str | Path) -> None:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(report, indent=2))
